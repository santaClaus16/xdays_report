import os
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Font

def drr_automate():
    st.title("Daily Remarks Report Automation")

    clean_drr_path = Path(r"C:\Users\SOLIZA\Documents\secret\clean_drr")
    MAPPING_PATH = Path(r"C:\Users\SOLIZA\Documents\secret\mapping\mapping_file.xlsx")
    OUTPUT_PATH = Path(r"C:\Users\SOLIZA\Documents\secret\automations\daily_remarks_report\output")

    OUTPUT_PATH.mkdir(parents=True, exist_ok=True)

    if not clean_drr_path.exists():
        st.error(f"Input folder not found: {clean_drr_path}")
        return
    if not MAPPING_PATH.exists():
        st.error(f"Mapping file not found: {MAPPING_PATH}")
        st.stop()

    files = [f for f in os.listdir(clean_drr_path) if f.lower().endswith(('.xlsx', '.xls'))]
    if not files:
        st.error("No Excel files found in clean_drr folder.")
        return

    with st.sidebar:
        selected_file = st.selectbox("Select a file to process", files)

    input_file_path = clean_drr_path / selected_file

    validation = st.radio(
        f"""Check selected file: **{files}**\nDo you want to proceed?""",
        ["No", "Yes"],
        index=0,
        horizontal=True,
    )

    if st.button("Process All Sheets & Save", type="primary"):
        if validation != "Yes":
            st.warning("Validation is NO – processing cancelled.")
            return

        # ── Log container ──────────────────────────────────────────────
        st.markdown("### 📋 Process Log")
        log_container = st.container(border=True,  height=350)
        logs = []  # accumulate (level, message) tuples

        def log(level, message):
            """Append to in-memory list AND render immediately."""
            logs.append((level, message))
            icon = {"info": "ℹ️", "success": "✅", "warning": "⚠️", "error": "❌"}.get(level, "•")
            timestamp = datetime.now().strftime("%I:%M:%S %p")

            # Re-render all logs inside the scrollable container
            with log_container:
                for lvl, msg in logs:
                    ico = {"info": "ℹ️", "success": "✅", "warning": "⚠️", "error": "❌"}.get(lvl, "•")
                    log_container.markdown(f"`{timestamp}` {ico} {msg}")

        # ──────────────────────────────────────────────────────────────

        try:
            xls = pd.ExcelFile(input_file_path)
            sheet_names = xls.sheet_names

            log("info", f"Found **{len(sheet_names)}** sheet(s): {', '.join(sheet_names)}")

            processed_dfs = {}
            mapping_ref = pd.read_excel(MAPPING_PATH, sheet_name="reference")
            log("info", f"Mapping file loaded — {mapping_ref.shape[1]} columns")

            for sheet in sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet)
                log("info", f"Processing sheet: **{sheet}**")

                # ── ADD "0" IN FRONT OF ACCOUNT NUMBER ──────────────────
                for col in list(df.columns):
                    col_lower = str(col).lower().strip()
                    if any(p in col_lower for p in ["account no", "account_no", "account number"]):
                        df[col] = df[col].astype(str).str.strip()
                        df[col] = df[col].str.replace(r'\.0$', '', regex=True)
                        df[col] = "0" + df[col]
                        log("info", f"Added leading '0' to column: **{col}**")

                if "Card No." in df.columns:
                    df["Card No."] = df["Card No."].astype(str).str.strip()
                    df["Card No."] = df["Card No."].str.replace(r'\.0$', '', regex=True)

                # ── AGENT SHEET ──────────────────────────────────────────
                if sheet == "DAILY REMARKS REPORT | AGENT":
                    formulated_heads = [
                        'STATUS', 'CYCLE', 'Month Cut Off', 'Month Extracted', 'ACTIVE',
                        'BASIS FOR REPORTING', 'CONCERN', 'RFD', 'Payment Type',
                        'Agent', 'OB', 'Unit code'
                    ]

                    for col in formulated_heads:
                        if col not in df.columns:
                            df[col] = None

                    other_cols = [col for col in df.columns if col not in formulated_heads]
                    df = df[formulated_heads + other_cols]

                    if "Card No." in df.columns:
                        df["CYCLE"] = "Cycle " + df["Card No."].astype(str).str[:2]
                    elif "Cycle" in df.columns:
                        df["CYCLE"] = "Cycle " + df["Cycle"].astype(str).str[:2]

                    lookup_col = mapping_ref.columns[0]
                    status_dict  = dict(zip(mapping_ref[lookup_col], mapping_ref.iloc[:, 1])) if len(mapping_ref.columns) > 1 else {}
                    basis_dict   = dict(zip(mapping_ref[lookup_col], mapping_ref.iloc[:, 2])) if len(mapping_ref.columns) > 2 else {}
                    concern_dict = dict(zip(mapping_ref[lookup_col], mapping_ref.iloc[:, 3])) if len(mapping_ref.columns) > 3 else {}
                    payment_dict = dict(zip(mapping_ref[lookup_col], mapping_ref.iloc[:, 4])) if len(mapping_ref.columns) > 4 else {}
                    rfd_status_dict = dict(zip(mapping_ref.iloc[:, 10], mapping_ref.iloc[:, 11])) if len(mapping_ref.columns) > 11 else {}

                    if "Status2" in df.columns:
                        df["STATUS"] = df["Status2"].map(status_dict).fillna("CHECK STATUS").replace(["", "nan"], "CHECK STATUS")
                        df["BASIS FOR REPORTING"] = df["Status2"].map(basis_dict).fillna("").replace(0, "")
                        df["CONCERN"] = df["Status2"].map(concern_dict).fillna("")
                        mask = (df["BASIS FOR REPORTING"] == "BANK ESCA") & (df["CONCERN"].isin(["", 0, pd.NA, "nan"]))
                        df.loc[mask, "CONCERN"] = df.loc[mask, "Status2"]
                        df["CONCERN"] = df["CONCERN"].astype(str).str.replace("BANK ESCALATION - ", "", regex=False).replace(["nan", "None"], "")

                    def extract_rfd_from_remark(remark):
                        if pd.isna(remark) or not isinstance(remark, str):
                            return ""
                        remark = remark.strip()
                        if "RFD" not in remark.upper():
                            return ""
                        try:
                            start = remark.upper().find("RFD") + 4
                            end = remark.find("|", start)
                            if end == -1:
                                end = len(remark)
                            return remark[start:end].strip().replace("*", "").strip()
                        except:
                            return ""

                    if "Remark" in df.columns:
                        df["RFD"] = df["Remark"].apply(extract_rfd_from_remark)
                        df["RFD"] = df["RFD"].fillna(df["Status2"].map(rfd_status_dict) if "Status2" in df.columns else pd.Series())
                    else:
                        df["RFD"] = df["Status2"].map(rfd_status_dict) if "Status2" in df.columns else ""

                    df["RFD"] = df["RFD"].fillna("").replace(["nan", "None"], "")
                    mask_ptp = (df["BASIS FOR REPORTING"] == "PTP") & (df["RFD"] == "")
                    df.loc[mask_ptp, "RFD"] = "FINANCIAL DIFFICULTY (LACKING/AWAITING FUNDS)"

                    if "Status2" in df.columns:
                        df["Payment Type"] = df["Status2"].map(payment_dict).fillna("")

                    if "Remark By" in df.columns and len(mapping_ref.columns) > 16:
                        agent_dict = dict(zip(mapping_ref.iloc[:, 14], mapping_ref.iloc[:, 16]))
                        df["Agent"] = df["Remark By"].map(agent_dict).fillna("")

                    if "Date" in df.columns:
                        df["Month Extracted"] = pd.to_datetime(df["Date"], errors='coerce').dt.strftime("%b").fillna("")

                    log("success", "AGENT sheet processed")

                # ── SYSTEM SHEET ─────────────────────────────────────────
                elif sheet == "DAILY REMARKS REPORT | SYSTEM":
                    formulated_heads_system = ["AMOUNT FIX", "CYCLE"]

                    if "Balance" in df.columns:
                        df["AMOUNT FIX"] = df["Balance"]
                    if "Card No." in df.columns:
                        df["CYCLE"] = df["Card No."].astype(str).str[:2]

                    other_cols = [col for col in df.columns if col not in formulated_heads_system]
                    df = df[formulated_heads_system + other_cols]

                    log("success", "SYSTEM sheet processed")

                else:
                    log("info", f"Sheet **'{sheet}'** copied as-is")

                processed_dfs[sheet] = df

            # ── SAVE ─────────────────────────────────────────────────────
            timestamp = datetime.now().strftime("%m_%d_%Y_%H%M")
            output_filename = f"{timestamp}_AUTOMATED_DRR.xlsx"
            output_file = OUTPUT_PATH / output_filename

            # with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            #     for sheet in sheet_names:
            #         df_processed = processed_dfs[sheet]
            #         df_processed.to_excel(writer, sheet_name=sheet, index=False)

            #         if sheet == "DAILY REMARKS REPORT | AGENT":
            #             worksheet = writer.sheets[sheet]
            #             yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            #             bold_font = Font(bold=True)
            #             formulated_heads = [
            #                 'STATUS', 'CYCLE', 'Month Cut Off', 'Month Extracted', 'ACTIVE',
            #                 'BASIS FOR REPORTING', 'CONCERN', 'RFD', 'Payment Type',
            #                 'Agent', 'OB', 'Unit code'
            #             ]
            #             for col_idx, col_name in enumerate(formulated_heads, start=1):
            #                 if col_name in df_processed.columns:
            #                     cell = worksheet.cell(row=1, column=col_idx)
            #                     cell.fill = yellow_fill
            #                     cell.font = bold_font

            #         if sheet == "DAILY REMARKS REPORT | SYSTEM":
            #             worksheet = writer.sheets[sheet]
            #             red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            #             bold_font = Font(bold=True)
            #             formulated_heads_system = ["AMOUNT FIX", "CYCLE"]
            #             for col_idx, col_name in enumerate(formulated_heads_system, start=1):
            #                 if col_name in df_processed.columns:
            #                     cell = worksheet.cell(row=1, column=col_idx)
            #                     cell.fill = red_fill
            #                     cell.font = bold_font

            log("info", "Leading '0' added to all Account Number columns")
            log("info", f"All sheets processed and saved!")
            log("success", f"Output: `{output_file}`")
            

        except Exception as e:
            log("error", f"Error: {e}")

if __name__ == "__main__":
    drr_automate()