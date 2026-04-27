import streamlit as st
import pandas as pd
import msoffcrypto
import re
from io import BytesIO
from datetime import datetime
import pyodbc
import os
from dotenv import load_dotenv

load_dotenv()

from utils import (
    load_mapping_file,
    apply_header_mapping,
    save_to_folder,
    clean_account_number_column,
    format_date_columns,
    force_columns_to_str,
)

MAPPING_PATH_WL = r"C:\Users\SPM\Desktop\eod_report\mapping\mapping_file_wl.xlsx"

def decrypt_file(file_bytes, password):
    """Helper function to decrypt msoffcrypto file"""
    try:
        decrypted = BytesIO()
        office_file = msoffcrypto.OfficeFile(BytesIO(file_bytes))
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
        decrypted.seek(0)
        return decrypted, None
    except msoffcrypto.exceptions.DecryptionError:
        return None, "Wrong password"
    except Exception as e:
        return None, str(e)

# ====================== DEFAULT SHEETS ======================
sheet_default = ["ActiveQry", "Active_updt", "Pulled Out", "!!DNC!!"]

# ====================== CLEAN COLUMN NAMES ======================
def clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Fix mixed-type column names warning"""
    new_cols = []
    for col in df.columns:
        col_str = str(col).strip()
        col_str = re.sub(r'[\r\n\t]', ' ', col_str)
        col_str = re.sub(r'\s+', ' ', col_str)
        new_cols.append(col_str)
    
    df = df.copy()
    df.columns = new_cols
    return df

# ====================== SAFE DATAFRAME VALIDATOR ======================
def safe_df(df, name):
    """Defensive check for DataFrame validity"""
    if df is None:
        st.error(f"❌ Missing sheet: **{name}**")
        return None
    if not isinstance(df, pd.DataFrame):
        st.error(f"❌ **{name}** is not a DataFrame (type: {type(df)})")
        return None
    return df

# ====================== CACHED TEMPLATE LOADER ======================
@st.cache_data(show_spinner="Loading template sheets...", ttl=600)
def load_template_sheets(template_file):
    """Cached function to load only default sheets from template"""
    if template_file is None:
        return {sheet: None for sheet in sheet_default}
    
    SHEET_DICT = {sheet: None for sheet in sheet_default}
    
    try:
        with st.spinner("Reading template sheets..."):
            excel_file = pd.ExcelFile(template_file)
            available_sheets = set(excel_file.sheet_names)

            for sheet_name in sheet_default:
                if sheet_name in available_sheets:
                    header_row = 1 if sheet_name in ["Active_updt", "Pulled Out"] else 0
                    
                    df = pd.read_excel(
                        template_file,
                        sheet_name=sheet_name,
                        header=header_row,
                        dtype=str,
                        engine="openpyxl"
                    )
                    
                    df = clean_column_names(df)
                    SHEET_DICT[sheet_name] = df
                    
                else:
                    st.warning(f"⚠️ Sheet '**{sheet_name}**' not found in template.")
        
        return SHEET_DICT
        
    except Exception as e:
        st.error(f"❌ Failed to load template: {e}")
        return {sheet: None for sheet in sheet_default}

# ====================== MAIN APP ======================
st.title("🔄 Worklist Processor")

# 1. Template Uploader
upload_file_template = st.file_uploader(
    "Upload your worklist template here",
    type=["xlsx"],
    key="template_uploader"
)

# Load template with caching
SHEET_DICT = load_template_sheets(upload_file_template)

if upload_file_template is None:
    st.info("👆 Please upload the **worklist template** first.")
    st.stop()

# Summary
loaded_count = sum(1 for v in SHEET_DICT.values() if v is not None)
st.success(f"✅ Template loaded successfully — {loaded_count}/{len(sheet_default)} default sheets ready")

# Safe DataFrame extraction with validation
df_active = safe_df(SHEET_DICT.get("ActiveQry"), "ActiveQry")
df_active_updt = safe_df(SHEET_DICT.get("Active_updt"), "Active_updt")
df_pulled_out = safe_df(SHEET_DICT.get("Pulled Out"), "Pulled Out")
df_dnc = safe_df(SHEET_DICT.get("!!DNC!!"), "!!DNC!!")

if df_active is None:
    st.error("❌ Critical: 'ActiveQry' sheet is missing from template!")
    st.stop()

# ====================== LOAD MAPPING RULES (once) ======================
try:
    rules = load_mapping_file(MAPPING_PATH_WL)
    if not isinstance(rules, dict):
        st.error(f"❌ Mapping file returned non-dict: {type(rules)}")
        rules = {"mapping": {}}
    standard_order = rules.get("standard_order", [])
except Exception as e:
    st.error(f"❌ Failed to load mapping rules: {e}")
    rules = {"mapping": {}}
    standard_order = []

# ====================== WORKLIST FILES UPLOADER ======================
uploaded_files = st.file_uploader(
    "Upload Worklist Files (c9, c14, c15, etc.)",
    type=["xlsx"],
    accept_multiple_files=True,
    key="worklist_uploader"
)

if uploaded_files:
    results = []
    all_dfs = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()

    for i, uploaded_file in enumerate(uploaded_files):
        filename = uploaded_file.name
        status_text.text(f"Processing: {filename} ({i+1}/{len(uploaded_files)})")

        match = re.search(r'[cC](\d+)', filename)
        if not match:
            st.error(f"❌ Could not find cycle number in: **{filename}**")
            continue

        c_number = match.group(1)
        auto_password = f"CYCLE_{c_number}*"

        with st.container(border=True):
            with st.expander(f"📁 Processing: {filename}", expanded=True):
                st.subheader(f"📄 {filename}")
                status = st.container()

                file_bytes = uploaded_file.read()
                decrypted_io = None

                # Auto password attempt
                with status:
                    st.info(f"🔄 Trying automatic password: `{auto_password}`")

                decrypted_io, error = decrypt_file(file_bytes, auto_password)

                if decrypted_io is None:
                    if "wrong password" in str(error).lower():
                        st.toast("❌ Wrong Password!", icon="🚫")
                        st.error(f"❌ Wrong password for **{filename}**")
                    else:
                        st.error(f"❌ Decryption failed: {error}")

                    # Manual password input
                    st.markdown("### 🔓 Enter Correct Password Manually")
                    manual_password = st.text_input(
                        "Enter password:",
                        value="",
                        type="password",
                        key=f"manual_pass_{i}_{filename}"
                    )

                    col1, col2, col3 = st.columns([2, 1, 1])
                    retry_btn = col1.button("🔓 Decrypt with this Password", key=f"retry_{i}")
                    skip_btn = col2.button("⏭️ Skip this file", key=f"skip_{i}")

                    if skip_btn:
                        st.info(f"⏭️ Skipped: {filename}")
                        continue

                    if retry_btn:
                        if not manual_password.strip():
                            st.error("Please enter a password")
                            continue
                        
                        decrypted_io, error = decrypt_file(file_bytes, manual_password)
                        if decrypted_io:
                            st.success(f"✅ Decrypted **{filename}** successfully with manual password!")
                        else:
                            st.error("❌ Password still incorrect. Please try again.")
                            continue
                else:
                    st.success(f"✅ Auto password worked for **{filename}**!")

                # Process decrypted file
                if decrypted_io is None:
                    continue

                try:
                    df = pd.read_excel(decrypted_io, dtype=str, engine="openpyxl")
                    df = clean_column_names(df)

                    with status:
                        st.success(f"✅ Loaded — {df.shape[0]:,} rows × {df.shape[1]} columns")

                    # Processing pipeline with defensive tuple handling
                    WL_STR_COLS = ["CUST_ID", "OFFICE_PH", "HOME_PH", "MOBILE_NO", "OB", "BOS", "AOD", "MAD", "PDA", "LPA", "PTP_AMT"]
                    result = force_columns_to_str(df, WL_STR_COLS)
                    df = result[0] if isinstance(result, tuple) else result

                    WL_DATE_COLS = ["LAST_PAYMENT_DATE", "PTP_DATE", "BIRTHDATE", "LAST_DUE_DATE", "D_CUST_OPN", "Birthdate"]
                    result = format_date_columns(df, date_cols=WL_DATE_COLS)
                    df = result[0] if isinstance(result, tuple) else result

                    result = clean_account_number_column(df, show_success=(i == 0), st=st)
                    df = result[0] if isinstance(result, tuple) else result

                    # Apply header mapping with defensive handling
                    result = apply_header_mapping(df, rules.get("mapping", {}), standard_order)
                    df = result[0] if isinstance(result, tuple) else result

                    # Collection Cycle
                    cycle_value = c_number
                    if "COLLECTION_CYCLE" in df.columns:
                        df["COLLECTION_CYCLE"] = (
                            df["COLLECTION_CYCLE"]
                            .fillna(cycle_value)
                            .astype(str)
                            .str.upper()
                            .str.replace(r"\s+", "_", regex=True)
                        )
                    else:
                        df["COLLECTION_CYCLE"] = cycle_value
                        st.warning(f"Created missing column COLLECTION_CYCLE = {cycle_value}")

                    with status:
                        st.success(f"✅ Mapping & processing completed for {filename}")

                    with st.expander("🔍 Preview first 10 rows", expanded=False):
                        st.dataframe(df.head(10), use_container_width=True, hide_index=True)

                    results.append({
                        "filename": filename,
                        "dataframe": df,
                        "c_number": c_number,
                    })
                    all_dfs.append(df)

                except Exception as e:
                    st.error(f"❌ Error processing {filename}: {str(e)}")
                    st.exception(e)  # Show full traceback

        # Update progress
        progress_bar.progress((i + 1) / len(uploaded_files))

    st.success(f"🎉 Finished processing all {len(uploaded_files)} worklist files!")

    # ====================== DOWNLOAD SECTION ======================
    if results:
        st.markdown("---")
        st.header("💾 Download Results")

        # Combined file download
        if len(results) > 1:
            combined_df = pd.concat(all_dfs, ignore_index=True)
            combined_filename = f"Combined_Worklists_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for result in results:
                    result["dataframe"].to_excel(writer, sheet_name=result["c_number"][:31], index=False)
                combined_df.to_excel(writer, sheet_name='ALL', index=False)
            output.seek(0)
            
            st.download_button(
                label="📥 Download Combined Excel",
                data=output.getvalue(),
                file_name=combined_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # Target Template Upload for OVERALL_CYCLE
        st.markdown("---")
        st.subheader("📝 Paste Combined Data to Template")
        st.info("Upload a template containing an 'OVERALL_CYCLE' sheet. The system will map and paste the combined data matching its headers.")
        
        overall_template = st.file_uploader(
            "Upload Output Template",
            type=["xlsx"],
            key="overall_template_uploader"
        )

        if overall_template:
            with st.spinner("Pasting data to template..."):
                try:
                    import openpyxl
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    # Ensure we have the combined data
                    combined_df = pd.concat(all_dfs, ignore_index=True)
                    
                    wb = openpyxl.load_workbook(overall_template)
                    if "OVERALL_CYCLE" not in wb.sheetnames:
                        st.error("❌ The uploaded template does not contain an 'OVERALL_CYCLE' sheet.")
                    else:
                        ws = wb["OVERALL_CYCLE"]
                        
                        # Get headers from first row
                        headers = [cell.value for cell in ws[1]]
                        
                        # Filter out None headers
                        valid_headers = [h for h in headers if h is not None]
                        
                        # Create mapped DataFrame based on valid_headers
                        mapped_data = {}
                        for header in valid_headers:
                            if header in combined_df.columns:
                                mapped_data[header] = combined_df[header].tolist()
                            else:
                                mapped_data[header] = [None] * len(combined_df)
                        
                        mapped_df = pd.DataFrame(mapped_data)
                        
                        # Delete existing data rows (keep header)
                        if ws.max_row > 1:
                            ws.delete_rows(2, ws.max_row)
                            
                        # Append new data
                        for r in dataframe_to_rows(mapped_df, index=False, header=False):
                            ws.append(r)
                            
                        # Save modified workbook to memory
                        output_template = BytesIO()
                        wb.save(output_template)
                        output_template.seek(0)
                        
                        st.success("✅ Successfully mapped and pasted data to 'OVERALL_CYCLE'!")
                        
                        st.download_button(
                            label="📥 Download Updated Template",
                            data=output_template.getvalue(),
                            file_name=f"Updated_Template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_updated_template"
                        )
                except Exception as e:
                    st.error(f"❌ Failed to process template: {e}")

        st.markdown("---")
        st.subheader("📄 Individual Downloads")
        # Individual downloads
        for result in results:
            csv_buffer = BytesIO()
            result["dataframe"].to_csv(csv_buffer, index=False)
            csv_buffer.seek(0)
            
            st.download_button(
                label=f"📥 Download {result['filename'][:20]}...",
                data=csv_buffer.getvalue(),
                file_name=f"{result['c_number']}_processed.csv",
                mime="text/csv",
                key=f"dl_{result['filename']}"
            )

        # Summary stats
        st.subheader("📊 Processing Summary")
        summary_data = []
        for result in results:
            summary_data.append({
                "File": result["filename"][:30],
                "Cycle": result["c_number"],
                "Rows": len(result["dataframe"]),
                "Columns": len(result["dataframe"].columns)
            })
        
        st.dataframe(pd.DataFrame(summary_data), use_container_width=True)

        # Save to folder option (if utils.save_to_folder exists)
        if 'save_to_folder' in globals():
            save_path = st.text_input("📁 Save folder path:", value=r"C:\Users\SPM\Desktop\eod_report\processed")
            if st.button("💾 Save All to Folder"):
                try:
                    os.makedirs(save_path, exist_ok=True)
                    for result in results:
                        filepath = os.path.join(save_path, f"{result['c_number']}_processed.xlsx")
                        result["dataframe"].to_excel(filepath, index=False)
                    st.success(f"✅ Saved {len(results)} files to {save_path}")
                except Exception as e:
                    st.error(f"❌ Save failed: {e}")

else:
    st.info("👆 Upload worklist files to start processing.")

st.markdown("---")
st.caption("✅ Code completed with error handling, downloads, and summary features.")